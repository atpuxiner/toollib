"""
@author axiner
@version v1.0.0
@created 2022/1/23 20:36
@abstract xlsx（基于openpyxl库）
@description
@history
"""
import typing as t

from toollib.validator import choicer

try:
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise

__all__ = [
    'ws_inserts',
    'ws_rows_value',
    'ws_cols_value',
    'ws_styles',
]


def ws_inserts(ws: Worksheet, values: t.List[list], index: int = None,
               mode: str = 'r', is_new: bool = True):
    """
    插入数据
    使用示例：
        from openpyxl import load_workbook
        wb = load_workbook('D:/tmp/t.xlsx')
        ws = wb.active
        values = [[1, 2, 3, 4, 5]]
        xlsx.ws_inserts(ws, values, index=1)
        .....
        +++++[更多详见参数或源码]+++++
    :param ws: Worksheet实例（openpyxl库）
    :param values: 值（eg: [[1, 2, 3], [4, 5, 6]]）
    :param index: 从哪开始插入值，若为None则追加（正整数）
    :param mode: 插入模式（r: 插入行值，c: 插入列值）
    :param is_new: 是否插入新的行|列（True: 是，False: 否，若有值则会覆盖）
    :return:
    """
    _len = len(values)
    mode = choicer(mode, choices=['r', 'c'], lable='mode')
    if index is None and mode == 'c':
        index = ws.max_column + 1
    if index is not None:
        if isinstance(index, int):
            if index < 1:
                raise ValueError('"index" greater than 0')
        else:
            raise ValueError('"index" only supported: int')
        if is_new is True:
            if mode == 'r':
                ws.insert_rows(index, _len)
            else:
                ws.insert_cols(index, _len)
        for r, item in enumerate(values):
            _i = r + index
            for c, v in enumerate(item, 1):
                if mode == 'r':
                    _r, _c = _i, c
                else:
                    _r, _c = c, _i
                ws.cell(_r, _c, v)
    else:
        for item in values:
            ws.append(item)


def ws_rows_value(ws: Worksheet):
    """
    所有行的值
    使用示例：
        from openpyxl import load_workbook
        wb = load_workbook('D:/tmp/t.xlsx')
        ws = wb.active
        rows = xlsx.ws_rows_value()
        # res: 返回所有行的值（生成器对象）
        +++++[更多详见参数或源码]+++++
    :param ws: Worksheet实例（openpyxl库）
    :return:
    """
    for item in ws.rows:
        yield [r.value for r in item]


def ws_cols_value(ws: Worksheet):
    """
    所有列的值
    使用示例：
        from openpyxl import load_workbook
        wb = load_workbook('D:/tmp/t.xlsx')
        ws = wb.active
        cols = xlsx.ws_cols_value()
        # res: 返回所有列的值（生成器对象）
        +++++[更多详见参数或源码]+++++
    :param ws: Worksheet实例（openpyxl库）
    :return:
    """
    for item in ws.columns:
        yield [c.value for c in item]


def ws_styles(
        ws,
        styles: dict,
        by_icells: str = None,
        by_cells: t.Union[str, t.List[str]] = None,
        by_rows: t.Union[int, t.List[int]] = None,
        by_cols: t.Union[str, t.List[str]] = None,
        exclude_icell: t.Union[t.List[t.Tuple[int, int]], t.Tuple[int, int]] = None,
):
    """
    修改样式
    使用示例：
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        wb = load_workbook('D:/tmp/t.xlsx')
        ws = wb.active
        fill = PatternFill(fill_type=None, start_color=’FFFFFF‘, end_color=‘000000’)
        styles = {'fill': fill}
        xlsx.ws_styles(styles, by_icells='0,0')
        .....
        +++++[更多详见参数或源码]+++++
    注：by_icells |by_cells |by_rows |by_cols，只支持其中一种方式（若传入多则只取第一种方式）
    :param ws: Worksheet实例（openpyxl库）
    :param styles: 样式。以字典形式传入（key可为：font, fill, border等）
    :param by_icells: 按单元格索引（若索引为0，则表示从1至最大行或列）。eg: '1:2,1:2' >>> 左行：1至2行，右列：1至2列
    :param by_cells: 按单元格。eg: 'A1' or ['A1', 'B1', 'C1']
    :param by_rows: 按行。eg: 1 or [1, 2, 3]
    :param by_cols: 按列。eg: 'A' or ['A', 'B', 'C']
    :param exclude_icell: 排除的单元格（by_icells方式下生效）
    :return:
    """
    if by_icells:
        by_icells = by_icells.strip()
        if by_icells in ['0', '0:0']:
            rows, cols = (1, ws.max_row), (1, ws.max_column)
        else:
            if by_icells.find(',') == -1:
                by_icells += ','
            rows, cols = by_icells.split(',')
            if not rows:
                rows = (1, ws.max_row)
            else:
                if rows.find(':') == -1:
                    _row = int(rows)
                    rows = (_row, _row)
                else:
                    _lrow, _rrow = rows.split(':')
                    rows = (int(_lrow), int(_rrow))
            if not cols:
                cols = (1, ws.max_column)
            else:
                if cols.find(':') == -1:
                    _col = int(cols)
                    cols = (_col, _col)
                else:
                    _lcol, _rcol = cols.split(':')
                    cols = (int(_lcol), int(_rcol))
        for r in range(rows[0], rows[1]+1):
            for c in range(cols[0], cols[1]+1):
                _curr_icell = (r, c)
                if isinstance(exclude_icell, tuple):
                    if _curr_icell == exclude_icell:
                        continue
                elif isinstance(exclude_icell, list):
                    if exclude_icell.count(_curr_icell) > 0:
                        continue
                ws_scope = ws.cell(r, c)
                for stl, stl_value in styles.items():
                    try:
                        getattr(ws_scope, stl)
                    except AttributeError:
                        raise
                    setattr(ws_scope, stl, stl_value)
    elif by_cells:
        if isinstance(by_cells, str):
            by_cells = [by_cells]
        for cell in by_cells:
            ws_scope = ws[cell]
            for stl, stl_value in styles.items():
                try:
                    getattr(ws_scope, stl)
                except AttributeError:
                    raise
                setattr(ws_scope, stl, stl_value)
    elif by_rows:
        if isinstance(by_rows, int):
            by_rows = [by_rows]
        for row in by_rows:
            ws_scope = ws.row_dimensions[row]
            for stl, stl_value in styles.items():
                try:
                    getattr(ws_scope, stl)
                except AttributeError:
                    raise
                setattr(ws_scope, stl, stl_value)
    elif by_cols:
        if isinstance(by_cols, str):
            by_cols = [by_cols]
        for col in by_cols:
            ws_scope = ws.column_dimensions[col]
            for stl, stl_value in styles.items():
                try:
                    getattr(ws_scope, stl)
                except AttributeError:
                    raise
                setattr(ws_scope, stl, stl_value)
    else:
        raise ValueError('"by_icells" |"by_cells" |"by_rows" |"by_cols" is required')
