import warnings
from pathlib import Path
from typing import Generator, Literal
from openpyxl import load_workbook, Workbook


def split_xlsx(
        filepath: str,
        max_rows: int,
        max_files: int = None,
        output_dir: str = None,
        part_sep: str = "_",
        part_prefix: str = "",
        part_zfill: int = 3,
        part_pos: Literal["after", "before"] = "after",
        sheet_name: int | str = 0,
) -> Generator[str, None, None]:
    """
    分割 xlsx 文件

    e.g.::

        for p in utils.split_xlsx(r'E:\tmp.xlsx'):
            print(p)

        +++++[更多详见参数或源码]+++++

    :param filepath: 文件路径
    :param max_rows: 最大行数
    :param max_files: 最大文件数
    :param output_dir: 输出目录
    :param part_sep: part分隔符
    :param part_prefix: part前缀
    :param part_zfill: part补零数
    :param part_pos: part编号位置
    :param sheet_name: 工作表
    :yields:
    """
    if max_rows <= 0:
        raise ValueError("max_rows must be a positive integer.")
    if max_files is not None and max_files <= 0:
        raise ValueError("max_files must be a positive integer or None.")

    src_path = Path(filepath)
    if not src_path.is_file() or src_path.suffix.lower() != ".xlsx":
        raise FileNotFoundError(f"Valid .xlsx file not found: {filepath}")

    output_dir = Path(output_dir) if output_dir else src_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    wb_src = load_workbook(str(src_path), read_only=True, data_only=True)
    all_sheet_names = wb_src.sheetnames
    if isinstance(sheet_name, int):
        if not (0 <= sheet_name < len(all_sheet_names)):
            raise IndexError(f"Sheet index {sheet_name} out of range. Available: 0–{len(all_sheet_names) - 1}")
        target_sheet_name = all_sheet_names[sheet_name]
    elif isinstance(sheet_name, str):
        if sheet_name not in all_sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {all_sheet_names}")
        target_sheet_name = sheet_name
    else:
        raise TypeError("sheet_name must be a string (sheet name) or an integer (sheet index).")
    ws_src = wb_src[target_sheet_name]
    rows_iter = ws_src.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        warnings.warn("No rows found in the sheet.")
        wb_src.close()
        return

    file_index = 0
    out_wb = None
    out_ws = None
    out_path = None
    written_rows = 0
    for row in rows_iter:
        if out_wb is None:
            if max_files is not None and file_index >= max_files:
                break
            part_name = f"{part_prefix}{str(file_index + 1).zfill(part_zfill)}"
            if part_pos == "after":
                file_name = f"{src_path.stem}{part_sep}{part_name}{src_path.suffix}"
            else:
                file_name = f"{part_name}{part_sep}{src_path.stem}{src_path.suffix}"
            out_path = output_dir / file_name
            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.title = target_sheet_name
            out_ws.append(header)
            written_rows = 0
        out_ws.append(row)
        written_rows += 1
        if written_rows >= max_rows:
            out_wb.save(out_path)
            out_wb.close()
            yield str(out_path)
            out_wb = None
            out_ws = None
            out_path = None
            file_index += 1

    if out_wb is not None:
        out_wb.save(out_path)
        out_wb.close()
        yield str(out_path)

    wb_src.close()
