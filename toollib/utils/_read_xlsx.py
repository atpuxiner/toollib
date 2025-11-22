import warnings
from typing import Generator
from openpyxl import load_workbook


def read_xlsx(
        filepath: str,
        column_names: list[str] = None,
        min_rows: int = None,
        max_rows: int = None,
        sheet_name: int | str = 0,
) -> Generator[tuple[int, dict[str, object]], None, None]:
    """
    读取 xlsx 文件

    e.g.::

        for idx, row in read_xlsx(r'E:\tmp.xlsx'):
            print(idx, row)

    :param filepath: 文件路径
    :param column_names: 列名称
    :param min_rows: 最小行
    :param max_rows: 最大行
    :param sheet_name: 工作表
    :return:
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    all_sheet_names = wb.sheetnames
    if isinstance(sheet_name, int):
        if not (0 <= sheet_name < len(all_sheet_names)):
            raise IndexError(f"Sheet index {sheet_name} out of range. Available: 0–{len(all_sheet_names) - 1}")
        target_sheet_name = all_sheet_names[sheet_name]
    elif isinstance(sheet_name, str):
        if sheet_name not in all_sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {all_sheet_names}")
        target_sheet_name = sheet_name
    else:
        raise TypeError("sheet_name must be a string (sheet name) or an integer (sheet index).")

    ws = wb[target_sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        warnings.warn("No rows found in the sheet.")
        wb.close()
        return
    if not header_row or all(cell is None for cell in header_row):
        wb.close()
        return

    actual_headers = [str(h).strip() if h is not None else "" for h in header_row]
    final_columns: list[str] = column_names if column_names is not None else actual_headers
    header_to_index = {name: i for i, name in enumerate(actual_headers)}
    for idx, row_values in enumerate(rows_iter):
        if idx < (min_rows or 0) - 1:
            continue
        if max_rows is not None and idx >= max_rows:
            break
        output_row = {}
        for col_name in final_columns:
            if col_name in header_to_index:
                idx_in_row = header_to_index[col_name]
                if idx_in_row < len(row_values):
                    output_row[col_name] = row_values[idx_in_row]
                else:
                    output_row[col_name] = None
            else:
                output_row[col_name] = None
        yield idx, output_row
    wb.close()
